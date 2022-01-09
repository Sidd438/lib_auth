from datetime import timedelta
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from lib_app.forms import UploadFileForm
from lib_app.models import Book, Issue, Renew
from lib_auth.settings import EMAIL_HOST_PASSWORD, EMAIL_HOST_USER
import smtplib

from librarian.forms import BookForm

def send_mail(email,message):
    with smtplib.SMTP('smtp.gmail.com',587) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
                smtp.sendmail(EMAIL_HOST_USER,email, message)


def handle_uploaded_file(user, spreadsheet):
    fs = FileSystemStorage()
    fs.save(r'spreadsheet\bookdata.xlsx', spreadsheet)
    workbook = load_workbook(filename=r"media\spreadsheet\bookdata.xlsx")
    sheet = workbook.active
    # 0name-1image_link-2summary-3author-4genre-5isbn-6location
    for value in sheet.iter_rows(min_row=1, min_col=1, max_col=7, values_only=True):
        if not(value[5]):
            break
        elif(Book.objects.filter(isbn=value[5]).exists()):
            book = Book.objects.filter(isbn=value[5]).first()
            if(value[0]):
                book.name = value[0]
            if(value[1]):
                book.image_link = value[1]
            if(value[2]):
                book.summary = value[2]
            if(value[3]):
                book.author = value[3]
            if(value[4]):
                book.genre = value[4]
            if(value[6]):
                book.location = value[6]
            book.save()
        else:
            book = Book.objects.create(name=value[0], image_link=value[1], summary=value[2],
                                       author=value[3], genre=value[4], isbn=value[5], location=value[6])
            book.save()
            user.libdata.books_added += 1
            user.libdata.save()
    import shutil
    shutil.rmtree(r"media\spreadsheet")

def process_return(issue_id, merit):
    returned = Issue.objects.get(id=issue_id)
    returned.returned = False
    returned.active = False
    returned.save()
    user = returned.user
    user.profile.merit = (user.profile.merit *user.profile.returns+int(merit))/(user.profile.returns+1)
    user.profile.returns += 1
    user.save()

def process_renew_request(id, flag):
    renew = Renew.objects.get(id)
    if(flag):
        issued = renew.issue
        td = timedelta(days=int(renew.time))
        issued.due_date = issued.due_date + td
        issued.save()
        renew.delete()
    else:
        renew.delete()

def lib_data():
    ReturnedsA = Issue.objects.filter(returned=True)
    IssuesA = Issue.objects.filter(pending=True)
    RenewsA = Renew.objects.all()
    form = UploadFileForm()
    book_form = BookForm()
    books = Book.objects.all()
    context = {'books': books, 'issues': IssuesA, 'returneds': ReturnedsA,
               'renews': RenewsA, 'form': form, 'book_form': book_form}
    return context
