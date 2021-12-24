from django.shortcuts import render, redirect
from lib_app.models import Issue, Book, Renew, Spreadsheet
from lib_app.forms import UploadFileForm
from django.contrib.auth import authenticate, login, logout
from datetime import datetime, timedelta
from django.core.mail import send_mail
from lib_auth.settings import EMAIL_HOST_PASSWORD, EMAIL_HOST_USER
from librarian.forms import BookForm
import smtplib
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from librarian.models import Libdata
from django.contrib import messages


def logoutA(request):
    logout(request)
    return redirect('/librarian')


def logina(request):
    return render(request, 'liblogin.html')


def logging(request):
    user = request.user
    print(EMAIL_HOST_USER+EMAIL_HOST_PASSWORD)
    if(request.method == 'POST'):
        form = BookForm(request.POST)
        if(form.is_valid()):
            form.save()
            user.profile.lib_data.books_added += 1
            user.profile.lib_data.save()
    if(request.POST.get("password")):
        user = authenticate(username=request.POST.get(
            "name"), password=request.POST.get("password"))
        if(user):
            if(user.groups.filter(name="Librarians").exists()):
                if not(user.profile.lib_data):
                    lib_data = Libdata.objects.create()
                    user.profile.lib_data = lib_data
                login(request,user)
                return libInterface(request)
        return render(request, "error.html")
    elif(request.FILES):
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.user,request.FILES['file'])
    elif(request.POST.get('time')):
        time = request.POST.get('time')
        issue_id = request.POST.get('issue_id')
        try:
            record = Issue.objects.filter(id=issue_id).first()
            book = record.book
            dt = datetime.now()
            td = timedelta(days=int(time))
            my_date = dt + td
            record.due_date = my_date
            record.pending = False
            record.issued = True
            record.save()
            book.issues += 1
            book.available = False
            book.save()
            user.profile.lib_data.books_issued += 1
            user.profile.lib_data.save()
            print("sending")
            with smtplib.SMTP('smtp.gmail.com',587) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
                msg = "Subject: Books Issued\n\n You have been Issued "+ book.name + " for "+time+" day/s"
                smtp.sendmail(EMAIL_HOST_USER,'thefrogwhoseesall@gmail.com', msg)
        except:
            pass
    elif(request.POST.get('reason')):
        issue_id = request.POST.get('issue_id')
        record = Issue.objects.get(id=issue_id)
        try:
            record.pending = False
            record.denied = True
            record.reason = request.POST.get('reason')
            record.save()
        except:
            print('pop')
            pass
    elif(request.POST.get('merit')):
        issue_id = request.POST.get('issue_id')
        merit = request.POST.get('merit')
        try:
            returned = Issue.objects.get(id=issue_id)
            returned.returned = False
            returned.active = False
            returned.save()
            user = returned.user
            user.profile.merit = (user.profile.merit *user.profile.returns+int(merit))/(user.profile.returns+1)
            user.profile.returns += 1
            user.save()
        except:
            pass
    elif(request.POST.get('accept')):
        try:    
            renew = Renew.objects.get(id=request.POST.get('timed'))
            issued = renew.issue
            td = timedelta(days=int(renew.time))
            issued.due_date = issued.due_date + td
            issued.save()
            renew.delete()
        except:
            pass    
    elif(request.POST.get('decline')):
        try:    
            renew = Renew.objects.get(id=request.POST.get('timed'))
            renew.delete()
        except:
            pass    
    return libInterface(request)

def libInterface(request):
    current_user = request.user
    messages.info(request, 'Your password has been changed successfully!')
    if(current_user.is_anonymous):
        return redirect('/librarian')
    ReturnedsA = Issue.objects.filter(returned=True)
    IssuesA = Issue.objects.filter(pending=True)
    RenewsA = Renew.objects.all()
    form = UploadFileForm()
    book_form = BookForm()
    books = Book.objects.all()
    context = {'books':books,'issues': IssuesA, 'returneds':ReturnedsA, 'renews': RenewsA, 'form':form, 'book_form':book_form}
    return render(request, "libinterface.html", context)

def handle_uploaded_file(user,spreadsheet):
    fs = FileSystemStorage()
    fs.save(r'spreadsheet\bookdata.xlsx', spreadsheet)
    workbook = load_workbook(filename=r"media\spreadsheet\bookdata.xlsx")
    sheet = workbook.active
    # 0name-1image_link-2summary-3author-4genre-5isbn-6location
    for value in sheet.iter_rows(min_row=1,min_col=1,max_col=7,values_only=True):
        if not(value[5]):
            break
        elif(Book.objects.filter(isbn=value[5]).exists()):
            book = Book.objects.filter(isbn=value[5]).first()
            if(value[0]):
                book.name = value[0]
            if(value[1]):
                book.image_link = value[1]
            if(value[2]):
                book.summary = value[2]
            if(value[3]):
                book.author = value[3]
            if(value[4]):
                book.genre = value[4]
            if(value[6]):
                book.location = value[6]
            book.save()
        else:
            book = Book.objects.create(name=value[0], image_link = value[1], summary=value[2], author=value[3], genre=value[4], isbn=value[5], location=value[6])
            book.save()
            user.profile.lib_data.books_added +=1
            user.profile.lib_data.save()
        import shutil
        shutil.rmtree(r"media\spreadsheet")

def profile(request):
    current_user = request.user
    if(current_user.is_anonymous):
        return redirect('/librarian')
    return render(request, 'lib_profile.html',{'user':request.user})